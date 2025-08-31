from openai import OpenAI
import httpx
import pandas as pd
import os
import time
import re
from openpyxl import load_workbook
import sys
from model_config import get_model_config, review_with_llm

# 初始化模型配置
model_config = get_model_config()
print(f"✅ 已加载模型配置: {model_config.get_provider()} - {model_config.get_model_name()}")

def reviewer(prompt: str) -> str:
    """使用配置的大模型进行评审"""
    return review_with_llm(prompt, model_config)

def extract_valid_content(input_str):
    start_marker = "<think>"
    end_marker = "</think>"
    
    start_pos = input_str.find(start_marker)
    if start_pos == -1:
        return input_str
    
    end_pos = input_str.find(end_marker, start_pos + len(start_marker))
    if end_pos == -1:
        return input_str
    
    return (input_str[:start_pos] + input_str[end_pos + len(end_marker):]).strip('\n')

def safe_get_value(row, key, default='无'):
    """安全获取并格式化字段值"""
    value = row.get(key, None)
    
    if pd.isna(value) or value is None:
        return default
    
    try:
        str_value = str(value).strip()
        
        if key in ['派生理由', '接口原型', '需求描述', '测试建议', '注释']:
            return str_value.strip('"') if str_value.startswith('"') else str_value
        
        return str_value
    except AttributeError as e:
        print(f"类型转换警告: {key} 原始值={value}, 错误={str(e)}")
        return default

def safe_save_to_excel(df, filepath, sheet_name='Sheet1'):
    """
    修复版保存函数 - 解决属性设置问题
    """
    max_retries = 3
    retry_delay = 2
    
    for attempt in range(max_retries):
        try:
            if not os.path.exists(filepath):
                df.to_excel(filepath, index=False, sheet_name=sheet_name)
                return True
            
            # 修复方法：使用不同的方式创建ExcelWriter
            book = load_workbook(filepath)
            
            # 获取目标sheet的最后一行
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                startrow = sheet.max_row
            else:
                sheet = book.create_sheet(sheet_name)
                startrow = 0
            
            # 处理空表头情况
            if startrow == 1 and sheet.cell(1, 1).value is None:
                startrow = 0
            
            # 使用pandas的ExcelWriter但不设置writer.book
            with pd.ExcelWriter(
                filepath,
                engine='openpyxl',
                mode='a',
                if_sheet_exists='overlay'
            ) as writer:
                # 直接写入数据
                header = False if startrow > 0 else True
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    header=header,
                    startrow=startrow
                )
            
            return True
            
        except PermissionError:
            if attempt < max_retries - 1:
                print(f"⚠️ 文件被占用，{retry_delay}秒后重试 ({attempt+1}/{max_retries})...")
                time.sleep(retry_delay)
            else:
                print(f"❌ 文件持续被占用，请关闭后重试: {filepath}")
                return False
        except Exception as e:
            print(f"❌ 写入失败: {str(e)}")
            # 特殊处理属性错误
            if "can't set attribute" in str(e):
                print("⚠️ 检测到属性设置错误，尝试替代方案...")
                return alternative_save_method(df, filepath, sheet_name)
            return False
    return False

def alternative_save_method(df, filepath, sheet_name='Sheet1'):
    """
    替代保存方法 - 当标准方法失败时使用
    """
    try:
        if not os.path.exists(filepath):
            df.to_excel(filepath, index=False, sheet_name=sheet_name)
            return True
        
        # 方法1: 读取整个文件然后重新保存
        try:
            existing_df = pd.read_excel(filepath, sheet_name=sheet_name, engine='openpyxl')
            combined_df = pd.concat([existing_df, df], ignore_index=True)
            combined_df.to_excel(filepath, index=False, sheet_name=sheet_name)
            print("✅ 使用替代方法1保存成功")
            return True
        except:
            # 方法2: 追加到CSV作为备份
            csv_path = filepath.replace('.xlsx', '_backup.csv')
            df.to_csv(csv_path, mode='a', header=not os.path.exists(csv_path), index=False)
            print(f"⚠️ 无法追加到Excel，已保存到CSV备份: {csv_path}")
            return False
    except Exception as e:
        print(f"❌ 替代保存方法失败: {str(e)}")
        return False

def format_elapsed_time(seconds):
    """格式化耗时显示"""
    if seconds < 60:
        return f"{seconds:.1f}秒"
    elif seconds < 3600:
        return f"{seconds/60:.1f}分钟"
    else:
        return f"{seconds/3600:.1f}小时"

if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    requirements_path = os.path.join(script_dir, "requirements.xlsx")  # 待评审的需求集合
    output_path = os.path.join(script_dir, "评审结果-cot.xlsx")  # 评审结果
    prompt_file = os.path.join(script_dir, "prompt.txt")
    checklist_file = os.path.join(script_dir, "checklist.txt")

    # 文件存在性检查
    for path in [requirements_path, prompt_file]:
        if not os.path.exists(path):
            raise FileNotFoundError(f"必要文件缺失: {path}")

    # 读取需求文件
    df_requirements = pd.read_excel(requirements_path, engine='openpyxl')
    
    # 读取提示模板
    with open(prompt_file, 'r', encoding='utf-8') as f:
        base_prompt = f.read()

    # 读取检查单
    with open(checklist_file, 'r', encoding='utf-8') as f:
        checklist = f.read()

    base_prompt = base_prompt.replace("[CHECKLIST]", checklist)
    
    # 获取总需求数量
    total_requirements = len(df_requirements)
    if total_requirements == 0:
        print("警告：需求表格为空！")
        exit()

    # 初始化进度统计变量
    start_time = time.time()
    processed_count = 0
    success_count = 0
    failed_count = 0

    # 创建日志文件
    log_file = os.path.join(script_dir, "review_log.txt")
    with open(log_file, 'w', encoding='utf-8') as log:
        log.write(f"评审开始时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        log.write(f"总需求数量: {total_requirements}\n")
        log.write("-"*50 + "\n")

    # 结果列表
    all_results = []

    for idx, row in df_requirements.iterrows():
        processed_count += 1
        config_id = safe_get_value(row, '标识')
        
        # 进度显示
        elapsed_time = time.time() - start_time
        avg_time_per_item = elapsed_time / processed_count if processed_count > 0 else 0
        remaining_items = total_requirements - processed_count
        estimated_remaining = remaining_items * avg_time_per_item
        
        progress_info = (
            f"\n{'='*75}\n"
            f"【处理进度】{processed_count}/{total_requirements} | "
            f"成功率: {success_count}/{processed_count} ({success_count/processed_count*100:.1f}%) | "
            f"耗时: {format_elapsed_time(elapsed_time)} | "
            f"预估剩余: {format_elapsed_time(estimated_remaining)}\n"
            f"当前处理: {config_id}\n"
            f"{'='*75}\n"
        )
        
        print("\033[H\033[J")  # ANSI清屏码
        print(progress_info)

        # 构建需求字符串
        requirement = f"""**标识**
{config_id}
**标题**
{safe_get_value(row, '标题')}
**版本信息**
{safe_get_value(row, '版本信息')}
**需求类型**
{safe_get_value(row, '需求类型')}
**是否派生的需求**
{safe_get_value(row, '是否派生的需求')}
**派生理由**
{safe_get_value(row, '派生理由')}
**接口原型**
{safe_get_value(row, '接口原型')}
**需求描述**
{safe_get_value(row, '需求描述')}
**测试建议**
{safe_get_value(row, '测试建议')}
**注释**
{safe_get_value(row, '注释')}
"""

        # 构造完整提示
        full_prompt = base_prompt.replace("[REQUIREMENT]", requirement)
        
        # 调用评审服务
        max_retries = 3
        review_result = ""
        for retry in range(max_retries):
            try:
                review_result = extract_valid_content(reviewer(full_prompt))
                if "Error:" not in review_result and review_result.strip():
                    break
                else:
                    print(f"⚠️ 评审返回异常，重试中 ({retry+1}/{max_retries})...")
                    time.sleep(2)
            except Exception as e:
                review_result = f"❌ 评审服务异常: {str(e)}"
                if retry < max_retries - 1:
                    time.sleep(3)
        
        # 评审结果
        result = {
            '标识': config_id,
            '作者': safe_get_value(row, '作者'),
            '失败': len(re.findall(r'\b失败\b', review_result)),
            '不确定': len(re.findall(r'\b不确定\b', review_result)),
            '不适用': len(re.findall(r'\b不适用\b', review_result)),
            '通过': len(re.findall(r'\b通过\b', review_result)),
            '额外问题': len(re.findall(r'\b额外问题\b', review_result)),
            '评审结果': review_result
        }
        
        # 添加到结果列表
        all_results.append(result)
        
        # 记录日志
        with open(log_file, 'a', encoding='utf-8') as log:
            log.write(f"需求 {config_id} 处理完成\n")
            log.write(f"评审摘要: 失败={result['失败']}, 通过={result['通过']}, 额外问题={result['额外问题']}\n")
            log.write("-"*50 + "\n")
        
        # 每处理完一个需求就尝试保存
        temp_df = pd.DataFrame([result])
        
        # 尝试保存
        if safe_save_to_excel(temp_df, output_path):
            success_count += 1
        else:
            failed_count += 1
            print(f"❌ 保存失败: 需求 {config_id}")
            
            # 写入紧急备份
            backup_path = output_path.replace(".xlsx", "_紧急备份.csv")
            temp_df.to_csv(backup_path, mode='a', header=not os.path.exists(backup_path), index=False)
            print(f"⚠️ 已创建紧急备份: {backup_path}")

    # 最终保存所有结果
    try:
        final_df = pd.DataFrame(all_results)
        final_df.to_excel(output_path, index=False)
        print(f"✅ 最终结果已保存至: {output_path}")
    except Exception as e:
        print(f"❌ 最终保存失败: {str(e)}")
        # 尝试保存为CSV
        csv_path = output_path.replace('.xlsx', '.csv')
        pd.DataFrame(all_results).to_csv(csv_path, index=False)
        print(f"⚠️ 结果已保存为CSV: {csv_path}")

    # 最终统计
    total_time = time.time() - start_time
    print("\n" + "="*70)
    print(f"✅ 处理完成! 总计: {total_requirements} 条需求")
    print(f"✔️ 成功保存: {success_count} 条")
    print(f"❌ 失败保存: {failed_count} 条")
    print(f"⏱️ 总耗时: {format_elapsed_time(total_time)}")
    print(f"📊 平均速度: {total_time/total_requirements:.2f} 秒/条")
    print(f"📝 日志文件: {log_file}")
    print(f"💾 结果文件: {output_path}")
    print("="*70)
