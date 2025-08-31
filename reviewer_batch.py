from openai import OpenAI
import httpx
import pandas as pd
import os
import time
import re
from openpyxl import load_workbook
import sys
from model_config import get_model_config, review_with_llm

# 初始化模型配置
model_config = get_model_config()
print(f"✅ 已加载模型配置: {model_config.get_provider()} - {model_config.get_model_name()}")

def reviewer(prompt: str) -> str:
    """使用配置的大模型进行评审"""
    return review_with_llm(prompt, model_config)

def extract_valid_content(input_str):
    start_marker = "<think>"
    end_marker = "</think>"
    
    start_pos = input_str.find(start_marker)
    if start_pos == -1:
        return input_str
    
    end_pos = input_str.find(end_marker, start_pos + len(start_marker))
    if end_pos == -1:
        return input_str
    
    return (input_str[:start_pos] + input_str[end_pos + len(end_marker):]).strip('\n')

def safe_get_value(row, key, default='无'):
    """安全获取并格式化字段值"""
    value = row.get(key, None)
    
    if pd.isna(value) or value is None:
        return default
    
    try:
        str_value = str(value).strip()
        
        if key in ['派生理由', '接口原型', '需求描述', '测试建议', '注释']:
            return str_value.strip('"') if str_value.startswith('"') else str_value
        
        return str_value
    except AttributeError as e:
        print(f"类型转换警告: {key} 原始值={value}, 错误={str(e)}")
        return default

def safe_save_to_excel(df, filepath, sheet_name='Sheet1'):
    """
    修复版保存函数 - 解决属性设置问题
    """
    max_retries = 3
    retry_delay = 2
    
    for attempt in range(max_retries):
        try:
            if not os.path.exists(filepath):
                df.to_excel(filepath, index=False, sheet_name=sheet_name)
                return True
            
            # 修复方法：使用不同的方式创建ExcelWriter
            book = load_workbook(filepath)
            
            # 获取目标sheet的最后一行
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                startrow = sheet.max_row
            else:
                sheet = book.create_sheet(sheet_name)
                startrow = 0
            
            # 处理空表头情况
            if startrow == 1 and sheet.cell(1, 1).value is None:
                startrow = 0
            
            # 使用pandas的ExcelWriter但不设置writer.book
            with pd.ExcelWriter(
                filepath,
                engine='openpyxl',
                mode='a',
                if_sheet_exists='overlay'
            ) as writer:
                # 直接写入数据
                header = False if startrow > 0 else True
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    header=header,
                    startrow=startrow
                )
            
            return True
            
        except PermissionError:
            if attempt < max_retries - 1:
                print(f"⚠️ 文件被占用，{retry_delay}秒后重试 ({attempt+1}/{max_retries})...")
                time.sleep(retry_delay)
            else:
                print(f"❌ 文件持续被占用，请关闭后重试: {filepath}")
                return False
        except Exception as e:
            print(f"❌ 写入失败: {str(e)}")
            # 特殊处理属性错误
            if "can't set attribute" in str(e):
                print("⚠️ 检测到属性设置错误，尝试替代方案...")
                return alternative_save_method(df, filepath, sheet_name)
            return False
    return False

def alternative_save_method(df, filepath, sheet_name='Sheet1'):
    """
    替代保存方法 - 当标准方法失败时使用
    """
    try:
        if not os.path.exists(filepath):
            df.to_excel(filepath, index=False, sheet_name=sheet_name)
            return True
        
        # 方法1: 读取整个文件然后重新保存
        try:
            existing_df = pd.read_excel(filepath, sheet_name=sheet_name, engine='openpyxl')
            combined_df = pd.concat([existing_df, df], ignore_index=True)
            combined_df.to_excel(filepath, index=False, sheet_name=sheet_name)
            print("✅ 使用替代方法1保存成功")
            return True
        except:
            # 方法2: 追加到CSV作为备份
            csv_path = filepath.replace('.xlsx', '_backup.csv')
            df.to_csv(csv_path, mode='a', header=not os.path.exists(csv_path), index=False)
            print(f"⚠️ 无法追加到Excel，已保存到CSV备份: {csv_path}")
            return False
    except Exception as e:
        print(f"❌ 替代保存方法失败: {str(e)}")
        return False

def format_elapsed_time(seconds):
    """格式化耗时显示"""
    if seconds < 60:
        return f"{seconds:.1f}秒"
    elif seconds < 3600:
        return f"{seconds/60:.1f}分钟"
    else:
        return f"{seconds/3600:.1f}小时"

if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 文件夹路径配置
    interfaces_dir = os.path.join(script_dir, "接口需求集合")  # 存放所有接口需求Excel文件的文件夹
    results_dir = os.path.join(script_dir, "评审结果")        # 存放所有评审结果的文件夹
    prompt_file = os.path.join(script_dir, "prompt_batch.txt")
    checklist_file = os.path.join(script_dir, "checklist.txt")
    
    # 创建结果目录（如果不存在）
    os.makedirs(results_dir, exist_ok=True)
    
    # 文件存在性检查
    for path in [interfaces_dir, prompt_file, checklist_file]:
        if not os.path.exists(path):
            raise FileNotFoundError(f"必要文件/文件夹缺失: {path}")

    # 读取提示模板
    with open(prompt_file, 'r', encoding='utf-8') as f:
        base_prompt = f.read()

    # 读取检查单
    with open(checklist_file, 'r', encoding='utf-8') as f:
        checklist = f.read()

    base_prompt = base_prompt.replace("[CHECKLIST]", checklist)
    
    # 获取所有接口文件
    interface_files = [f for f in os.listdir(interfaces_dir) if f.endswith('.xlsx')]
    total_interfaces = len(interface_files)
    
    if total_interfaces == 0:
        print("警告：接口需求文件夹中没有Excel文件！")
        exit()

    # 初始化进度统计
    start_time = time.time()
    processed_count = 0
    success_count = 0
    failed_count = 0
    
    # 创建汇总结果列表
    summary_results = []
    
    # 创建主日志文件
    main_log_file = os.path.join(results_dir, "评审总日志.txt")
    with open(main_log_file, 'w', encoding='utf-8') as log:
        log.write(f"接口需求集合批量评审开始时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        log.write(f"接口总数: {total_interfaces}\n")
        log.write(f"接口列表: {', '.join(interface_files)}\n")
        log.write("="*80 + "\n\n")

    print(f"\n{'='*80}")
    print(f"开始批量处理接口需求集合")
    print(f"接口总数: {total_interfaces}")
    print(f"接口列表: {', '.join(interface_files)}")
    print(f"{'='*80}\n")

    # 处理每个接口文件
    for interface_file in interface_files:
        processed_count += 1
        interface_path = os.path.join(interfaces_dir, interface_file)
        interface_name = os.path.splitext(interface_file)[0]
        output_file = f"评审结果-{interface_name}.xlsx"
        output_path = os.path.join(results_dir, output_file)
        
        # 为每个接口创建单独的日志文件
        interface_log_file = os.path.join(results_dir, f"评审日志-{interface_name}.txt")
        
        print(f"\n{'='*80}")
        print(f"处理接口 {processed_count}/{total_interfaces}: {interface_name}")
        print(f"{'='*80}")
        
        # 读取需求文件
        try:
            df_requirements = pd.read_excel(interface_path, engine='openpyxl')
        except Exception as e:
            error_msg = f"❌ 读取接口文件失败: {interface_file}, 错误: {str(e)}"
            print(error_msg)
            with open(main_log_file, 'a', encoding='utf-8') as log:
                log.write(f"{error_msg}\n")
            failed_count += 1
            continue
        
        # 获取需求数量
        total_requirements = len(df_requirements)
        if total_requirements == 0:
            warning_msg = f"⚠️ 接口 {interface_name} 的需求表格为空！跳过处理。"
            print(warning_msg)
            with open(main_log_file, 'a', encoding='utf-8') as log:
                log.write(f"{warning_msg}\n")
            continue
        
        # 记录接口处理开始
        interface_start_time = time.time()
        with open(interface_log_file, 'w', encoding='utf-8') as log:
            log.write(f"接口需求集合评审开始时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            log.write(f"接口名称: {interface_name}\n")
            log.write(f"需求数量: {total_requirements}\n")
            log.write("-"*50 + "\n")
        
        # 构建接口需求集合字符串
        requirement_text = ""
        
        # 收集所有需求
        for idx, row in df_requirements.iterrows():
            config_id = safe_get_value(row, '标识')
            
            # 显示进度
            progress = f"正在收集需求 {idx+1}/{total_requirements}: {config_id}"
            print(f"\r{progress}", end="", flush=True)
            
            requirement_text += f"""\n**需求 {idx+1} - 标识: {config_id}**
**标题**
{safe_get_value(row, '标题')}
**版本信息**
{safe_get_value(row, '版本信息')}
**需求类型**
{safe_get_value(row, '需求类型')}
**是否派生的需求**
{safe_get_value(row, '是否派生的需求')}
**派生理由**
{safe_get_value(row, '派生理由')}
**接口原型**
{safe_get_value(row, '接口原型')}
**需求描述**
{safe_get_value(row, '需求描述')}
**测试建议**
{safe_get_value(row, '测试建议')}
**注释**
{safe_get_value(row, '注释')}
**作者**
{safe_get_value(row, '作者')}\n
"""
        
        print(f"\n\n✅ 需求收集完成，共 {total_requirements} 条需求")
        
        # 构造完整提示
        full_prompt = base_prompt.replace("[REQUIREMENT]", requirement_text)
        
        # 记录开始评审时间
        review_start_time = time.time()
        print("🚀 开始调用评审服务...")
        
        # 调用评审服务
        max_retries = 3
        review_result = ""
        for retry in range(max_retries):
            try:
                print(f"🔄 第 {retry+1} 次尝试调用评审服务...")
                review_result = extract_valid_content(reviewer(full_prompt))
                if "Error:" not in review_result and review_result.strip():
                    print("✅ 评审服务调用成功")
                    break
                else:
                    print(f"⚠️ 评审返回异常，{2}秒后重试 ({retry+1}/{max_retries})...")
                    time.sleep(2)
            except Exception as e:
                review_result = f"❌ 评审服务异常: {str(e)}"
                print(f"❌ 评审服务异常: {str(e)}")
                if retry < max_retries - 1:
                    time.sleep(3)
        
        # 计算评审耗时
        review_time = time.time() - review_start_time
        
        # 提取评审结果中的关键信息
        result_match = re.search(r'\[评审结果\](.*?)(?=\[\/评审结果\]|$)', review_result, re.DOTALL)
        if result_match:
            review_content = result_match.group(1).strip()
        else:
            review_content = review_result
        
        # 统计各类结果数量
        failure_count = len(re.findall(r'\b失败\b', review_content))
        uncertain_count = len(re.findall(r'\b不确定\b', review_content))
        not_applicable_count = len(re.findall(r'\b不适用\b', review_content))
        pass_count = len(re.findall(r'\b通过\b', review_content))
        extra_issues_count = len(re.findall(r'\b额外问题\b', review_content))
        
        # 创建评审结果记录
        result = {
            '接口名称': interface_name,
            '需求数量': total_requirements,
            '失败': failure_count,
            '不确定': uncertain_count,
            '不适用': not_applicable_count,
            '通过': pass_count,
            '额外问题': extra_issues_count,
            '评审耗时(秒)': round(review_time, 2),
            '评审结果': review_content
        }
        
        # 添加到汇总结果
        summary_results.append({
            '接口名称': interface_name,
            '需求数量': total_requirements,
            '失败': failure_count,
            '不确定': uncertain_count,
            '不适用': not_applicable_count,
            '通过': pass_count,
            '额外问题': extra_issues_count,
            '评审耗时(秒)': round(review_time, 2)
        })
        
        # 记录接口日志
        with open(interface_log_file, 'a', encoding='utf-8') as log:
            log.write(f"接口评审完成时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            log.write(f"评审耗时: {review_time:.2f}秒\n")
            log.write(f"评审摘要: 失败={failure_count}, 不确定={uncertain_count}, ")
            log.write(f"不适用={not_applicable_count}, 通过={pass_count}, 额外问题={extra_issues_count}\n")
            log.write("-"*50 + "\n")
            log.write(f"详细评审结果:\n{review_content}\n")
            log.write("="*80 + "\n")
        
        # 保存评审结果
        result_df = pd.DataFrame([result])
        
        # 尝试保存
        save_success = False
        for save_retry in range(3):
            try:
                if safe_save_to_excel(result_df, output_path):
                    save_success = True
                    break
                else:
                    print(f"保存失败，{2}秒后重试 ({save_retry+1}/3)...")
                    time.sleep(2)
            except Exception as e:
                print(f"保存异常: {str(e)}，{2}秒后重试...")
                time.sleep(2)
        
        if save_success:
            success_count += 1
            print(f"✅ 接口 {interface_name} 评审结果已保存")
        else:
            failed_count += 1
            print(f"❌ 接口 {interface_name} 保存失败，创建紧急备份...")
            # 写入紧急备份
            backup_path = output_path.replace(".xlsx", "_紧急备份.csv")
            result_df.to_csv(backup_path, index=False)
            print(f"⚠️ 已创建紧急备份: {backup_path}")
        
        # 计算接口处理耗时
        interface_time = time.time() - interface_start_time
        
        # 记录到主日志
        with open(main_log_file, 'a', encoding='utf-8') as log:
            log.write(f"接口 {interface_name} 处理完成\n")
            log.write(f"  需求数量: {total_requirements}\n")
            log.write(f"  处理耗时: {interface_time:.2f}秒\n")
            log.write(f"  评审结果: 失败={failure_count}, 不确定={uncertain_count}, ")
            log.write(f"不适用={not_applicable_count}, 通过={pass_count}, 额外问题={extra_issues_count}\n")
            log.write("-"*50 + "\n")
        
        # 显示接口处理摘要
        print(f"\n📋 接口 {interface_name} 处理完成!")
        print(f"⏱️ 处理耗时: {format_elapsed_time(interface_time)}")
        print(f"📈 评审结果:")
        print(f"   ❌ 失败: {failure_count}")
        print(f"   ❓ 不确定: {uncertain_count}")
        print(f"   ➖ 不适用: {not_applicable_count}")
        print(f"   ✅ 通过: {pass_count}")
        print(f"   ⚠️ 额外问题: {extra_issues_count}")
        print(f"📝 日志文件: {interface_log_file}")
        print(f"💾 结果文件: {output_path}")
        
        # 显示简短评审结果摘要
        print("\n📋 评审结果摘要:")
        print("-"*50)
        short_review = review_content[:500] + "..." if len(review_content) > 500 else review_content
        print(short_review)
        
        # 添加处理间隔，避免API调用过于频繁
        if processed_count < total_interfaces:
            print(f"\n⏳ 等待3秒后处理下一个接口...")
            time.sleep(3)

    # 保存汇总结果
    if summary_results:
        summary_df = pd.DataFrame(summary_results)
        summary_path = os.path.join(results_dir, "接口评审汇总.xlsx")
        summary_df.to_excel(summary_path, index=False)
        print(f"\n✅ 接口评审汇总已保存: {summary_path}")
    
    # 最终统计
    total_time = time.time() - start_time
    
    print("\n" + "="*80)
    print(f"✅ 批量处理完成!")
    print(f"📋 接口总数: {total_interfaces}")
    print(f"✔️ 成功处理: {success_count}")
    print(f"❌ 处理失败: {failed_count}")
    print(f"⏱️ 总耗时: {format_elapsed_time(total_time)}")
    print(f"📊 平均速度: {total_time/total_interfaces:.2f} 秒/接口")
    print(f"📝 主日志文件: {main_log_file}")
    print(f"💾 结果目录: {results_dir}")
    print("="*80)
    
    # 显示汇总统计
    if summary_results:
        print("\n📊 接口评审汇总:")
        print("-"*80)
        for result in summary_results:
            print(f"{result['接口名称']}: {result['需求数量']}需求, "
                  f"失败{result['失败']}, 通过{result['通过']}, "
                  f"耗时{result['评审耗时(秒)']}秒")
